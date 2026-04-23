"""
ProofServe Dashboard Build Script
Reads data from Google Sheets tab-by-tab via API, generates dashboard HTML.
Runs in GitHub Actions on a schedule.
Also supports local mode: python build.py --local path/to/spreadsheet.xlsx
"""

import os
import sys
import json
import re
import statistics
from collections import defaultdict
from pathlib import Path

# ═══════════════════════════════════════════
# CONFIG
# ═══════════════════════════════════════════

MONTHS = ['Jan', 'Feb', 'Mar', 'Apr']
MONTH_MAP = {'2026-01': 'Jan', '2026-02': 'Feb', '2026-03': 'Mar', '2026-04': 'Apr'}

WEIGHTS = {
    'Service': {'Chats': 30, '1st Attempt Breach (1+)': 25, 'Job SLA Breach': 20,
                'Bad Address': 10, 'Ad Hoc': 8, 'Attempt Limit': 5, 'Nudge Server (1+)': 2},
    'ESC Lead': {'Chats': 30, '1st Attempt Breach (1+)': 25, 'Job SLA Breach': 20,
                 'Bad Address': 10, 'Ad Hoc': 8, 'Attempt Limit': 5, 'Nudge Server (1+)': 2},
    'Off App': {'Chats': 30, '1st Attempt Breach (1+)': 25, 'Job SLA Breach': 20,
                'Bad Address': 10, 'Ad Hoc': 8, 'Attempt Limit': 5, 'Nudge Server (1+)': 2},
    'QA': {'Affidavit Preparation': 75, 'Affidavit Changes': 20, 'Ad Hoc': 5},
    'Dispatch': {'Ad Hoc': 20, 'Assign Server': 80},
}

ROLE_NORM = {
    'Escalation Leads': 'ESC Lead', 'Service': 'Service', 'QA': 'QA',
    'Dispatch': 'Dispatch', 'Off App': 'Off App', 'Team Lead': 'Team Lead'
}


# ═══════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════

def sf(v):
    if v is None or v == '' or v == 'None': return None
    if isinstance(v, (int, float)): return float(v)
    try: return float(str(v).replace(',', '').replace('%', '').strip())
    except: return None

def si(v):
    if v is None or v == '' or v == 'None': return 0
    if isinstance(v, (int, float)): return int(v)
    try: return int(float(str(v).replace(',', '').strip()))
    except: return 0

def pm(v):
    if v is None: return None
    s = str(v)
    m = re.search(r'(2026-0[1-4])', s)
    return MONTH_MAP.get(m.group(1)) if m else None

def cn(v):
    if not v or str(v).strip() == '': return None
    s = str(v).strip()
    if s.startswith('@'): s = s[1:].strip()
    if s.lower() in ('none', 'server', 'client', 'nan', ''): return None
    return s

def quality_score(wq, total_tasks):
    if total_tasks == 0: return None
    anchor = 0.03 + (75.0 / total_tasks)
    return round(max(0, min(100, 100 - (wq / anchor) * 100)), 1)


# ═══════════════════════════════════════════
# SHEET READER ABSTRACTION
# ═══════════════════════════════════════════

class SheetReader:
    """Abstract reader that works with both gspread (API) and openpyxl (local xlsx)"""
    
    def __init__(self, source, local_path=None):
        self.source = source  # 'api' or 'local'
        if source == 'local':
            import openpyxl
            self.wb = openpyxl.load_workbook(local_path, data_only=True)
        else:
            self.spreadsheet = local_path  # actually a gspread Spreadsheet object
        self._cache = {}
    
    def get_sheet(self, name):
        """Get all values from a sheet as a 2D list (0-indexed)"""
        if name in self._cache:
            return self._cache[name]
        
        if self.source == 'local':
            ws = self.wb[name]
            data = []
            for r in range(1, ws.max_row + 1):
                row = []
                for c in range(1, ws.max_column + 1):
                    row.append(ws.cell(r, c).value)
                data.append(row)
        else:
            try:
                ws = self.spreadsheet.worksheet(name)
                data = ws.get_all_values()
            except Exception as e:
                print(f"  Warning: Could not read sheet '{name}': {e}")
                data = []
        
        self._cache[name] = data
        print(f"  Read '{name}': {len(data)} rows x {len(data[0]) if data else 0} cols")
        return data
    
    def cell(self, sheet_name, row, col):
        """Get cell value (1-indexed like openpyxl)"""
        data = self.get_sheet(sheet_name)
        if row <= 0 or col <= 0 or row > len(data): return None
        r = data[row - 1]
        if col > len(r): return None
        v = r[col - 1]
        if v == '' or v is None: return None
        return v


# ═══════════════════════════════════════════
# DATA EXTRACTION
# ═══════════════════════════════════════════

def get_reader(local_path=None):
    if local_path:
        print(f"Reading local file: {local_path}")
        return SheetReader('local', local_path)
    
    import gspread
    from google.oauth2.service_account import Credentials
    
    creds_path = os.environ.get('GOOGLE_APPLICATION_CREDENTIALS')
    sheet_id = os.environ.get('SHEET_ID')
    
    if not creds_path or not sheet_id:
        print("ERROR: GOOGLE_APPLICATION_CREDENTIALS and SHEET_ID env vars required")
        sys.exit(1)
    
    scopes = [
        'https://www.googleapis.com/auth/spreadsheets.readonly',
        'https://www.googleapis.com/auth/drive.readonly'
    ]
    creds = Credentials.from_service_account_file(creds_path, scopes=scopes)
    gc = gspread.authorize(creds)
    
    print(f"Connecting to sheet {sheet_id}...")
    spreadsheet = gc.open_by_key(sheet_id)
    print(f"Connected: {spreadsheet.title}")
    
    return SheetReader('api', spreadsheet)


def extract_roster(sr):
    roster = {}
    data = sr.get_sheet('Roster1')
    for r in range(2, len(data)):  # 0-indexed, skip header rows
        row = data[r]
        for base in [0, 3, 6, 9, 12]:  # 0-indexed column groups
            if base + 2 >= len(row): continue
            t = row[base] if base < len(row) else None
            n = row[base + 1] if base + 1 < len(row) else None
            rl = row[base + 2] if base + 2 < len(row) else None
            if n and rl and t and str(n).strip() and str(rl).strip():
                nm = str(n).strip()
                role = ROLE_NORM.get(str(rl).strip(), str(rl).strip())
                roster[nm] = {'team': str(t).strip(), 'role': role}
    print(f"Roster: {len(roster)} reps")
    return roster


def detect_layout(data):
    if len(data) < 2 or len(data[0]) < 2: return 'tasks_first'
    return 'dates_first' if '2026' in str(data[0][1] or '') else 'tasks_first'


def parse_task_sheet(data):
    """Parse a Tasks or Adh sheet from raw 2D data"""
    if not data or len(data) < 4: return {}, []
    layout = detect_layout(data)
    task_row_idx = 1 if layout == 'dates_first' else 0  # 0-indexed
    month_row_idx = 0 if layout == 'dates_first' else 1
    
    cols = {}
    for c in range(1, len(data[0])):
        tn_raw = data[task_row_idx][c] if c < len(data[task_row_idx]) else None
        md_raw = data[month_row_idx][c] if c < len(data[month_row_idx]) else None
        if tn_raw and md_raw:
            mi = pm(md_raw)
            tn = str(tn_raw).strip()
            if mi: cols[(tn, mi)] = c
    
    task_names = sorted(set(k[0] for k in cols))
    records = {}
    for r in range(3, len(data)):  # Data starts at row 4 (0-indexed row 3)
        row = data[r]
        if not row or not row[0] or str(row[0]).strip() == '': break
        name = str(row[0]).strip()
        records[name] = {}
        for mi in MONTHS:
            records[name][mi] = {}
            for tn in task_names:
                c = cols.get((tn, mi))
                val = sf(row[c]) if c is not None and c < len(row) else None
                records[name][mi][tn] = val
    return records, task_names


def extract_fails(sr, roster):
    roster_names = set(roster.keys())
    all_fails = []
    
    # Affidavit Exceptions
    data = sr.get_sheet('Affidavit Exceptions')
    for r in range(1, len(data)):
        row = data[r]
        if len(row) < 15: continue
        month = pm(row[8] if len(row) > 8 else None)  # Col I (0-indexed 8)
        if not month: continue
        attrib = str(row[13] if len(row) > 13 else '').strip()  # Col N
        nm = cn(row[12] if len(row) > 12 else None)  # Col M
        nj = cn(row[9] if len(row) > 9 else None)   # Col J
        cat = str(row[1] if len(row) > 1 else '')    # Col B
        job = str(row[5] if len(row) > 5 else '')    # Col F
        note = str(row[10] if len(row) > 10 else '')[:200]  # Col K
        firm = cn(row[16] if len(row) > 16 else None)  # Col Q
        reason = str(row[11] if len(row) > 11 else '')[:100]  # Col L
        if attrib == 'Attributee' and nm:
            all_fails.append({'p': nm, 'm': month, 'b': 'Aff Exceptions', 'c': cat,
                              'j': job, 'note': note, 'f': firm, 'reason': reason})
        elif attrib == 'Both':
            if nm: all_fails.append({'p': nm, 'm': month, 'b': 'Aff Exceptions', 'c': cat,
                                     'j': job, 'note': note, 'f': firm, 'reason': reason})
            if nj and nj != nm: all_fails.append({'p': nj, 'm': month, 'b': 'Aff Exceptions', 'c': cat,
                                                  'j': job, 'note': note, 'f': firm, 'reason': reason})
    
    # Ops Sales Escal
    data = sr.get_sheet('Ops Sales Escal')
    for r in range(1, len(data)):
        row = data[r]
        if len(row) < 10: continue
        if str(row[8] if len(row) > 8 else '').strip().lower() != 'fail': continue  # Col I
        name = cn(row[6] if len(row) > 6 else None)  # Col G
        if not name: continue
        month = pm(row[9] if len(row) > 9 else None)  # Col J
        if not month: continue
        all_fails.append({'p': name, 'm': month, 'b': 'Ops Sales Escal',
                          'c': str(row[11] if len(row) > 11 else '')[:60],  # Col L
                          'j': str(row[1] if len(row) > 1 else ''),  # Col B
                          'note': str(row[12] if len(row) > 12 else '')[:200],  # Col M
                          'f': cn(row[14] if len(row) > 14 else None),  # Col O
                          'reason': ''})
    
    # Client Feedback
    data = sr.get_sheet('Client Feedback')
    for r in range(1, len(data)):
        row = data[r]
        if len(row) < 12: continue
        if str(row[11] if len(row) > 11 else '').strip().lower() != 'fail': continue  # Col L
        name = cn(row[6] if len(row) > 6 else None)  # Col G
        if not name: continue
        month = pm(row[8] if len(row) > 8 else None)  # Col I
        if not month: continue
        raw_b = str(row[1] if len(row) > 1 else '')  # Col B
        parts = re.split(r'\s*[-\u2013\u2014]\s*', raw_b)
        firm = parts[-1].strip() if len(parts) >= 2 else None
        all_fails.append({'p': name, 'm': month, 'b': 'Client Feedback', 'c': 'Client Feedback',
                          'j': str(row[12] if len(row) > 12 else ''),  # Col M
                          'note': str(row[9] if len(row) > 9 else '')[:200],  # Col J
                          'f': firm, 'reason': ''})
    
    # Ops Leadership
    data = sr.get_sheet('Ops Leadership')
    for r in range(1, len(data)):
        row = data[r]
        if len(row) < 13: continue
        if str(row[12] if len(row) > 12 else '').strip().lower() != 'fail': continue  # Col M
        name = cn(row[7] if len(row) > 7 else None)  # Col H
        if not name: continue
        month = pm(row[10] if len(row) > 10 else None)  # Col K
        if not month: month = pm(row[3] if len(row) > 3 else None)  # fallback Col D
        if not month: continue
        all_fails.append({'p': name, 'm': month, 'b': 'Ops Leadership',
                          'c': str(row[14] if len(row) > 14 else '')[:60],  # Col O
                          'j': str(row[5] if len(row) > 5 else ''),  # Col F
                          'note': str(row[15] if len(row) > 15 else '')[:200],  # Col P
                          'f': cn(row[6] if len(row) > 6 else None),  # Col G
                          'reason': str(row[14] if len(row) > 14 else '')[:100]})  # Col O
    
    all_fails = [f for f in all_fails if f['p'] in roster_names]
    print(f"Total fails (roster-filtered): {len(all_fails)}")
    return all_fails


def extract_rework(sr):
    data = sr.get_sheet('Rework Rate')
    rework = {}
    # Jan=Col7(idx6), Feb=Col13(idx12), Mar=Col19(idx18), Apr=Col25(idx24)
    for r in range(2, len(data)):
        row = data[r]
        if not row or not row[0] or str(row[0]).strip() == '': break
        n = str(row[0]).strip()
        rework[n] = {
            'Jan': sf(row[6]) if len(row) > 6 else 0,
            'Feb': sf(row[12]) if len(row) > 12 else 0,
            'Mar': sf(row[18]) if len(row) > 18 else 0,
            'Apr': sf(row[24]) if len(row) > 24 else 0,
        }
        for k in rework[n]:
            if rework[n][k] is None: rework[n][k] = 0
    return rework


def extract_tot(sr):
    data = sr.get_sheet('Time on Tasks')
    agg = defaultdict(lambda: defaultdict(lambda: {
        'tms': [], 'fms': [], 'paces': defaultdict(int),
        'overlaps': defaultdict(int), 'n': 0
    }))
    for r in range(1, len(data)):
        row = data[r]
        if len(row) < 29: continue
        person = cn(row[0])  # Col A
        if not person: continue
        mi = pm(row[3])  # Col D
        if not mi: continue
        tm = sf(row[21])  # Col V (0-indexed 21)
        fm = sf(row[19])  # Col T (0-indexed 19)
        pace = str(row[27] if len(row) > 27 else '').strip().lower()  # Col AB
        overlap = str(row[28] if len(row) > 28 else '').strip().lower()  # Col AC
        d = agg[person][mi]
        d['n'] += 1
        if tm is not None: d['tms'].append(tm)
        if fm is not None: d['fms'].append(fm)
        if pace: d['paces'][pace] += 1
        if overlap: d['overlaps'][overlap] += 1
    
    tot_data = {}
    for p, months in agg.items():
        tot_data[p] = {}
        for mi, d in months.items():
            if d['n'] == 0: continue
            tp = sum(d['paces'].values()) or 1
            to = sum(d['overlaps'].values()) or 1
            tot_data[p][mi] = {
                'med': round(statistics.median(d['tms']), 1) if d['tms'] else 0,
                'foc': round(sum(d['fms']) / max(sum(d['tms']), 0.01) * 100, 1) if d['tms'] else 0,
                'lng': round(d['paces'].get('long', 0) / tp * 100, 1),
                'rap': round(d['paces'].get('rapid', 0) / tp * 100, 1),
                'sol': round(d['overlaps'].get('solo', 0) / to * 100, 1),
                'tot': round(sum(d['tms']), 0), 'n': d['n']
            }
    print(f"ToT: {len(tot_data)} people with data")
    return tot_data


# ═══════════════════════════════════════════
# BUILD MONTHLY DATA
# ═══════════════════════════════════════════

def build_monthly_data(sr, roster, all_fails, rework, tot_data):
    roles_tasks = {}
    roles_adh = {}
    for role, ts, ads in [
        ('Service', 'Service (Tasks)', 'Service (Adh)'),
        ('ESC Lead', 'ESC Lead (Tasks)', 'ESC Lead (Adh)'),
        ('QA', 'QA (Tasks)', 'QA (Adh)'),
        ('Dispatch', 'Dispatch (Tasks)', 'Dispatch (Adh)'),
        ('Off App', 'Off App (Tasks)', 'Off App (Adh)')
    ]:
        t_data, _ = parse_task_sheet(sr.get_sheet(ts))
        a_data, _ = parse_task_sheet(sr.get_sheet(ads))
        roles_tasks[role] = t_data
        roles_adh[role] = a_data
        print(f"  {role}: {len(t_data)} reps in Tasks, {len(a_data)} in Adh")
    
    fail_map = defaultdict(lambda: defaultdict(list))
    for f in all_fails: fail_map[f['p']][f['m']].append(f)
    
    output = {'months': {}}
    for mi in MONTHS:
        md = {'roles': {}}
        for role in ['Service', 'ESC Lead', 'QA', 'Dispatch', 'Off App']:
            wts = WEIGHTS[role]; reps = []
            t_data = roles_tasks[role]; a_data = roles_adh[role]
            for name in set(list(t_data.keys()) + list(a_data.keys())):
                ri = roster.get(name, {})
                if ri.get('role') != role: continue
                tasks_m = t_data.get(name, {}).get(mi, {})
                adh_m = a_data.get(name, {}).get(mi, {})
                total_tasks = sum(v or 0 for v in tasks_m.values() if isinstance(v, (int, float)))
                if total_tasks == 0: continue
                wtotal = 0; wsum = 0; ps_dict = {}; tk_dict = {}
                for tname, wt in wts.items():
                    t_count = tasks_m.get(tname) or 0
                    adh_val = adh_m.get(tname)
                    if isinstance(t_count, (int, float)) and t_count > 0: tk_dict[tname] = int(t_count)
                    if adh_val is not None and isinstance(adh_val, (int, float)) and t_count and t_count > 0:
                        ps = min(100, max(0, adh_val * 100))
                        ps_dict[tname] = round(ps, 1)
                        wtotal += ps * wt / 100; wsum += wt
                if wsum == 0: continue
                wadh = round(wtotal * 100 / wsum, 1)
                if wadh < 5: continue
                person_fails = fail_map.get(name, {}).get(mi, [])
                mf = len(person_fails)
                wq = mf / total_tasks if mf > 0 else 0
                qs = quality_score(wq, total_tasks) if mf > 0 else 100.0
                ops = round(wadh * 0.4 + qs * 0.6, 1)
                rw = round((rework.get(name, {}).get(mi, 0) or 0) * 100, 2)
                tot_p = tot_data.get(name, {}).get(mi)
                rep = {'n': name, 't': ri['team'].split('(')[0].strip(), 'tf': ri['team'], 'r': role,
                       'w': wadh, 'q': qs, 'o': ops, 'tt': int(total_tasks), 'mf': mf,
                       'wq': round(wq * 100, 4), 'rr': rw, 'ps': ps_dict, 'tk': tk_dict,
                       'fd': [{'b': f['b'], 'c': f['c'], 'j': f['j'], 'note': f['note'],
                               'f': f.get('f', ''), 'reason': f.get('reason', ''), 'm': f['m']}
                              for f in person_fails[:10]]}
                if tot_p: rep['tot'] = tot_p
                reps.append(rep)
            reps.sort(key=lambda x: x['o'], reverse=True)
            md['roles'][role] = reps
        
        all_r = [r for reps in md['roles'].values() for r in reps]
        ranked = [r for r in all_r if r.get('o')]
        if ranked:
            md['org'] = {
                'avg_ops': round(sum(r['o'] for r in ranked) / len(ranked), 1),
                'avg_wadh': round(sum(r['w'] for r in ranked) / len(ranked), 1),
                'avg_qs': round(sum(r['q'] for r in ranked) / len(ranked), 1),
                'total_tasks': sum(r['tt'] for r in all_r),
                'total_fails': sum(r['mf'] for r in all_r),
                'ranked_count': len(ranked),
                'avg_rework': round(sum(r['rr'] for r in all_r) / max(len(all_r), 1), 2)
            }
        else:
            md['org'] = {'avg_ops': 0, 'avg_wadh': 0, 'avg_qs': 0, 'total_tasks': 0,
                         'total_fails': 0, 'ranked_count': 0, 'avg_rework': 0}
        
        teams = {}
        for r in ranked:
            tn = r['tf']
            if tn not in teams: teams[tn] = {'ops': [], 'qs': [], 'wadh': [], 'tasks': 0, 'errors': 0, 'reps': []}
            teams[tn]['ops'].append(r['o']); teams[tn]['qs'].append(r['q']); teams[tn]['wadh'].append(r['w'])
            teams[tn]['tasks'] += r['tt']; teams[tn]['errors'] += r['mf']
            teams[tn]['reps'].append({'name': r['n'], 'ops': r['o'], 'role': r['r']})
        md['teams'] = {tn: {'avg_ops': round(sum(v['ops']) / len(v['ops']), 1),
            'avg_qs': round(sum(v['qs']) / len(v['qs']), 1),
            'avg_wadh': round(sum(v['wadh']) / len(v['wadh']), 1),
            'n': len(v['ops']), 'tasks': v['tasks'], 'errors': v['errors'],
            'reps': sorted(v['reps'], key=lambda x: x['ops'], reverse=True)}
            for tn, v in sorted(teams.items(), key=lambda x: sum(x[1]['ops']) / len(x[1]['ops']), reverse=True)}
        
        md['winners'] = {}
        for role, reps in md['roles'].items():
            valid = [r for r in reps if r.get('o')]
            if valid: md['winners'][role] = {'name': valid[0]['n'], 'ops': valid[0]['o'], 'team': valid[0]['t']}
        
        output['months'][mi] = md
        print(f"{mi}: OPS={md['org']['avg_ops']} Fails={md['org']['total_fails']} Ranked={md['org']['ranked_count']}")
    
    # Firm data
    firm_data = defaultdict(lambda: {'count': 0, 'months': defaultdict(int), 'people': set(), 'details': []})
    for f in all_fails:
        firm = f.get('f') or 'Unknown / Not recorded'
        firm_data[firm]['count'] += 1; firm_data[firm]['months'][f['m']] += 1
        firm_data[firm]['people'].add(f['p'])
        if len(firm_data[firm]['details']) < 25:
            firm_data[firm]['details'].append({'b': f['b'], 'c': f['c'], 'j': f['j'], 'note': f['note'],
                'f': firm, 'person': f['p'], 'team': roster.get(f['p'], {}).get('team', ''),
                'reason': f.get('reason', ''), 'm': f['m']})
    output['firms'] = {k: {'c': v['count'], 'm': dict(v['months']), 'p': list(v['people']), 'd': v['details']}
        for k, v in sorted(firm_data.items(), key=lambda x: x[1]['count'], reverse=True)[:60]}
    
    return output


# ═══════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════

def main():
    local_path = None
    if len(sys.argv) > 1 and sys.argv[1] == '--local' and len(sys.argv) > 2:
        local_path = sys.argv[2]
    
    sr = get_reader(local_path)
    
    print("\n=== Extracting data ===")
    roster = extract_roster(sr)
    all_fails = extract_fails(sr, roster)
    rework = extract_rework(sr)
    tot_data = extract_tot(sr)
    
    print("\n=== Building monthly data ===")
    data = build_monthly_data(sr, roster, all_fails, rework, tot_data)
    
    data_json = json.dumps(data, separators=(',', ':'), default=str)
    print(f"\nData JSON: {len(data_json):,} chars")
    
    template_path = Path(__file__).parent / 'template.html'
    if not template_path.exists():
        print(f"ERROR: template.html not found at {template_path}")
        sys.exit(1)
    
    template = template_path.read_text(encoding='utf-8')
    html = template.replace('__DASHBOARD_DATA_PLACEHOLDER__', data_json)
    
    output_dir = Path(__file__).parent / 'output'
    output_dir.mkdir(exist_ok=True)
    output_file = output_dir / 'index.html'
    output_file.write_text(html, encoding='utf-8')
    
    print(f"\nDashboard written to {output_file} ({output_file.stat().st_size:,} bytes)")
    print("Done!")


if __name__ == '__main__':
    main()
